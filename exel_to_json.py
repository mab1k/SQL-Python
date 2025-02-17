import openpyxl
import json

days = ['ПОНЕДЕЛЬНИК', "ВТОРНИК", 'СРЕДА', 'ЧЕТВЕРГ', 'ПЯТНИЦА', 'СУББОТА']
type_of_activity = ['ЛК', 'ПР', 'СР']
auditorium = ['ауд.', 'физ.', 'комп.']
group_temp = ['КМБО-02-23', 'КМБО-05-23', 'КМБО-02-22', 'КМБО-05-22', 'КМБО-05-21', 'КМБО-02-21', 'КМБО-02-20',
              'КМБО-05-20', "КММО-01-23", "КММО-01-22", "КММО-02-22"]
is_none = [None, None]
const_include_days = 4
const_not_include_days = 5
day_week = 14

'''
    Получаем список, который сожержит все дни недели, нач. и оконч. пар, номер пары и четность недели
'''


def get_main(book):
    sheet = book.active
    count = 0
    result = list()
    for row in sheet.iter_rows(min_row=4, max_row=87, min_col=1, max_col=5, values_only=True):
        temp = ''
        for cel in row:
            if cel is not None:
                temp += cel + ' '
        lst = temp.split()
        count += 1
        result.append(lst)
    return result


'''
    Получаем список нужных нам групп, в котором содержатся:
    1. Индексы начала (откуда считывать)
    2. Индексы конца (до куда считывать)
    3. Названия групп

    Возвращает список с группами и выше перечисленными параметрами 
'''


def get_groups(book):
    sheet = book.active
    groups = list()
    count_group_id1 = 0

    for row in sheet.iter_rows(min_row=2, max_row=2, min_col=2, max_col=sheet.max_column, values_only=True):
        for cel in row:
            count_group_id1 += 1
            if cel in group_temp:
                temp = list()
                temp.append(count_group_id1 + 1)
                temp.append(count_group_id1 + 4)
                temp.append(cel)

                groups.append(temp)
    return groups




'''
    Обрабатываем курс
'''


def parse(sheet, groups, kurs_work, book):
    main_info = get_main(book)

    for el in groups:
        count = 0
        kurs_work[el[2]] = list()
        group_name = el[2]
        id_begin = el[0]
        id_end = el[1]

        temp_day = ''

        for row, el_main_info in zip(sheet.iter_rows(min_row=4, max_row=87, min_col=id_begin, max_col=id_end), main_info):
            temp_str = ''
            for cel in row:
                if cel.value is not None:
                    temp_str += cel.value + ' '

            lst = temp_str.split()
            lst = el_main_info + lst
            temp_dict = dict()

            if count == 0:
                for el in days:
                    if el in lst:
                        temp_day = el
            temp_dict["День недели"] = temp_day

            if count == 0:
                temp_dict["Номер пары"] = lst[1]
                temp_dict["Нач. занятия"] = lst[2]
                temp_dict["Оконч. занятия"] = lst[3]
                temp_dict["Неделя занятия"] = lst[4]
            else:
                temp_dict["Номер пары"] = lst[0]
                temp_dict["Нач. занятия"] = lst[1]
                temp_dict["Оконч. занятия"] = lst[2]
                temp_dict["Неделя занятия"] = lst[3]

            count += 1
            if count == day_week:
                count = 0

            if len(lst) == const_not_include_days or len(lst) == const_include_days:
                temp_dict["Дисциплина занятия"] = ""
                temp_dict["Вид занятия"] = ""
                temp_dict["ФИО преподавателя"] = ""
                temp_dict["Номер аудитории"] = ""
            else:
                indices_name = [i for i, x in enumerate(lst) if x in type_of_activity]
                indices_values_name = [x for i, x in enumerate(lst) if x in type_of_activity]
                indices_auditorium = [i for i, x in enumerate(lst) if x in auditorium]
                if len(indices_name) == 1:
                    if count != 1:
                        discipline_lesson = lst[4:indices_name[0]]
                        discipline_lesson_str = ' '.join(discipline_lesson)
                        temp_dict["Дисциплина занятия"] = discipline_lesson_str
                    else:
                        discipline_lesson = lst[5:indices_name[0]]
                        discipline_lesson_str = ' '.join(discipline_lesson)
                        temp_dict["Дисциплина занятия"] = discipline_lesson_str
                    if not indices_values_name:
                        temp_dict["Вид занятия"] = ""
                    else:
                        temp_dict["Вид занятия"] = indices_values_name[0]

                    if not indices_auditorium:
                        temp_dict["ФИО преподавателя"] = ""
                        temp_dict["Номер аудитории"] = ""
                    else:
                        full_name = lst[(indices_name[0] + 1):indices_auditorium[0]]
                        temp_dict["ФИО преподавателя"] = ' '.join(full_name)

                        audience_number = lst[indices_auditorium[0]:]
                        audience_number_str = ' '.join(audience_number)

                        temp_dict["Номер аудитории"] = audience_number_str

                elif len(indices_name) == 2:

                    if count != 1:
                        discipline_lesson = lst[4:indices_name[0]]
                        discipline_lesson_str = ' '.join(discipline_lesson)
                        temp_dict["Дисциплина занятия"] = discipline_lesson_str
                    else:
                        discipline_lesson = lst[5:indices_name[0]]
                        discipline_lesson_str = ' '.join(discipline_lesson)
                        temp_dict["Дисциплина занятия"] = discipline_lesson_str

                    if not indices_values_name:
                        temp_dict["Вид занятия"] = ""
                    else:
                        temp_dict["Вид занятия"] = ', '.join(indices_values_name)

                    if not indices_auditorium:
                        temp_dict["ФИО преподавателя"] = ""
                        temp_dict["Номер аудитории"] = ""
                    else:
                        full_name = lst[(indices_name[1] + 1):indices_auditorium[0]]
                        temp_dict["ФИО преподавателя"] = ' '.join(full_name)

                        audience_number = lst[indices_auditorium[0]:]
                        audience_number_str = ' '.join(audience_number)
                        temp_dict["Номер аудитории"] = audience_number_str

            kurs_work[group_name].append(temp_dict)


'''
    Считывание и запись в эксель
'''


def parse_exel(kurs, book):
    sheet = book.active

    groups = get_groups(book)

    kurs_work = dict()

    parse(sheet, groups, kurs_work, book)

    with open(f"kurs_work-{kurs}.json", "w", encoding='utf-8') as outfile:
        json.dump(kurs_work, outfile, ensure_ascii=False, indent=4)


if __name__ == '__main__':
    book = openpyxl.load_workbook("1-kurs.xlsx", read_only=True)
    parse_exel(1, book)
    book = openpyxl.load_workbook("2-kurs.xlsx", read_only=True)
    parse_exel(2, book)
    book = openpyxl.load_workbook("3-kurs.xlsx", read_only=True)
    parse_exel(3, book)
    book = openpyxl.load_workbook("4-kurs.xlsx", read_only=True)
    parse_exel(4, book)
    book = openpyxl.load_workbook("mag-1-kurs.xlsx", read_only=True)
    parse_exel(5, book)
    book = openpyxl.load_workbook("mag-2-kurs.xlsx", read_only=True)
    parse_exel(6, book)
